import openpyxl
from pynput.keyboard import Key, Controller

def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_row)

def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_column)

def readData(file,sheetName,rownum,columno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row = rownum, column = columno).value

def writeData(file,sheetName,rownum,columno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]

def capturaScreen(captura):
    keyboard = Controller()
    # Captura de pantalla está habilitado (1) entonces toma screenshots
    if captura == 1:
        keyboard.press(Key.print_screen)
